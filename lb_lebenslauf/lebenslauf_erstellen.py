import sqlite3
import os
from datetime import datetime
import tkinter as tk
import sys
from tkinter import ttk, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


# Funktion zur Herstellung der Verbindung
def create_connection():
    db_path = '/home/birgit/PycharmProjects/Projekt//lebenslauf.db' #Hinweis: Geben Sie den vollständigen, absoluten Pfad zur Datei an.
    conn = None
    try:
        conn = sqlite3.connect(db_path)
        print("Verbindung zur Datenbank erfolgreich!")
    except sqlite3.Error as e:
        print(f"Fehler beim Verbinden mit der Datenbank: {e}")
    return conn

# Verbindung aufbauen
conn = create_connection()
cursor = conn.cursor()



# Hilfsfunktionen
def safe_value(val):
    return "" if val is None or (isinstance(val, str) and val.strip() == "") else str(val)

def format_datum(v):
    return v.strftime("%d.%m.%Y") if hasattr(v, "strftime") else safe_value(v)

def is_valid_leihfirma(text):
    return bool(text) and text.strip().lower() not in ("none", "null", "-")

def get_bewerbungen():
    try:
        conn = create_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT b.bwg_id,
                   bw.bw_vorname || ' ' || bw.bw_nachname || ' - ' || f.f_name
            FROM tbl_bewerbung b
            JOIN tbl_bewerber bw ON bw.bw_id = b.bwg_bw_id
            JOIN tbl_firma f ON f.f_id = b.bwg_f_id
            ORDER BY bw.bw_nachname, bw.bw_vorname
        """)
        return cur.fetchall()
    except Exception as e:
        messagebox.showerror("Fehler", f"Datenbankfehler: {e}")
        return []
    finally:
        conn.close()

# Formatvorlagen
font_title = Font(name="Arial", bold=True, size=20)
font_section = Font(name="Arial", bold=True, size=12)
font_bold = Font(name="Arial", bold=True)
font_normal = Font(name="Arial")
align_left = Alignment(horizontal="left")
align_center = Alignment(horizontal="center")
align_indent = Alignment(indent=1)

def add_section(ws, title, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=title)
    c.font = font_section
    c.alignment = align_left
    for col in range(1, 5):
        ws.cell(row, col).border = Border(bottom=Side(style="medium"))
    ws.row_dimensions[row].height = 20
    return row + 2

def auto_adjust_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 5, 50)

def export_lebenslauf(bwg_id):
    try:
        conn = create_connection()
        cur = conn.cursor()
    except Exception as e:
        messagebox.showerror("Fehler", f"DB-Verbindung: {e}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Lebenslauf"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    for m in ('left', 'right', 'top', 'bottom'):
        setattr(ws.page_margins, m, 0.75 if m in ('top', 'bottom') else 0.5)

    row = 1
    ws.cell(row, 1, "Lebenslauf").font = font_title
    row += 1  # Leerzeile

    # Foto in C1
    img = os.path.join(os.path.dirname(__file__), "foto.png")
    if os.path.exists(img):
        im = Image(img)
        im.width = im.height = 120
        im.anchor = "C1"
        ws.add_image(im)

    # Persönliche Daten
    cur.execute("""
        SELECT bw.bw_vorname || ' ' || bw.bw_nachname, bw.bw_geburtsdatum,
               bw.bw_strasse, bw.bw_plz, bw.bw_ort, bw.bw_telefon, bw.bw_mail
        FROM tbl_bewerber bw
        JOIN tbl_bewerbung b ON b.bwg_bw_id = bw.bw_id
        WHERE b.bwg_id = ?
    """, (bwg_id,))
    pdata = cur.fetchone()
    if not pdata:
        messagebox.showerror("Fehler", f"Keine Daten für ID {bwg_id}")
        conn.close()
        return

    ws.cell(row, 1, "Name").font = font_bold
    ws.cell(row, 2, safe_value(pdata[0])); row += 1
    ws.cell(row, 1, "Geburtsdatum").font = font_bold
    ws.cell(row, 2, f"Geboren am {format_datum(pdata[1])} in {pdata[4]}"); row += 1
    ws.cell(row, 1, "Straße").font = font_bold
    ws.cell(row, 2, safe_value(pdata[2])); row += 1

    for label, value in [("Ort", f"{pdata[3]} {pdata[4]}"), ("Telefon", safe_value(pdata[5])), ("E‑Mail", safe_value(pdata[6]))]:
        ws.cell(row, 1, label).font = font_bold
        cell = ws.cell(row, 2, value)
        if label == "E‑Mail" and value:
            cell.hyperlink = f"mailto:{value}"
            cell.font = Font(name="Arial", color="0000FF", underline="single")
        row += 1

    row += 1

    # Berufserfahrung
    row = add_section(ws, "Berufserfahrung", row)
    cur.execute("""
        SELECT ag.ag_name, ag.ag_datum_von, ag.ag_datum_bis, ag.ag_zeit,
               ag.ag_funktion, ag.ag_leihfirma, bag.bwg_ag_id
        FROM tbl_bwg_ag bag
        JOIN tbl_arbeitgeber ag ON ag.ag_id = bag.bwg_ag_ag_id
        WHERE bag.bwg_ag_bwg_id = ?
        ORDER BY ag.ag_datum_bis DESC
    """, (bwg_id,))
    for name, von, bis, zeit, funktion, leihfirma, ag_id in cur.fetchall():
        period = f"{format_datum(von)} – {format_datum(bis) if bis else 'heute'}"
        info = name
        if is_valid_leihfirma(leihfirma):
            info += f" / {leihfirma}"
        ws.cell(row, 1, period).font = font_normal
        ws.cell(row, 2, info).font = font_bold
        ws.cell(row, 3, safe_value(zeit)).font = font_normal
        row += 1

        if funktion and funktion.lower() != "none":
            ws.cell(row, 2, safe_value(funktion)).font = font_normal
            row += 1

        cur.execute("""
            SELECT t.t_name
            FROM tbl_bwg_ag_t bagt
            JOIN tbl_taetigkeit t ON t.t_id = bagt.bwg_ag_t_t_id
            WHERE bagt.bwg_ag_t_bwg_ag_id = ?
            ORDER BY t.t_name
        """, (ag_id,))
        for (tn,) in cur.fetchall():
            ws.cell(row, 2, f"• {safe_value(tn)}").alignment = align_indent
            row += 1
        row += 1

    # Ausbildung
    row = add_section(ws, "Ausbildung", row)
    cur.execute("""
        SELECT ab.ab_datum_von, ab.ab_datum_bis,
               ab.ab_name_staette, ab.ab_name_ausbildung,
               ab.ab_abschluss,
               GROUP_CONCAT(swp.ab_swp_name, ', ')
        FROM tbl_bwg_ab bab
        JOIN tbl_ausbildung ab ON ab.ab_id = bab.bwg_ab_ab_id
        LEFT JOIN tbl_bwg_ab_swp bas ON bas.bwg_ab_swp_bwg_ab_id = bab.bwg_ab_id
        LEFT JOIN tbl_ab_schwerpunkt swp ON swp.ab_swp_id = bas.bwg_ab_swp_ab_swp_id
        WHERE bab.bwg_ab_bwg_id = ?
        GROUP BY ab.ab_id
        ORDER BY ab.ab_datum_von DESC
    """, (bwg_id,))
    for von, bis, inst, aus, abs_, swp in cur.fetchall():
        ws.cell(row, 1, f"{format_datum(von)} - {format_datum(bis) if bis else 'heute'}").font = font_normal
        ws.cell(row, 2, inst).font = font_bold
        row += 1
        aus_text = aus
        if swp:
            sp = ", ".join(s.strip() for s in swp.split(",") if s.lower() != "(keine)")
            if sp:
                aus_text += f" ({sp})"
        if abs_:
            aus_text += f" | Abschluss: {abs_}"
        ws.cell(row, 2, aus_text)
        row += 2

    # Kenntnisse
    row = add_section(ws, "Kenntnisse", row)
    cur.execute("""
        SELECT k.k_name, k.k_stufe
        FROM tbl_bwg_k bwgk
        JOIN tbl_kenntnisse k ON k.k_id = bwgk.bwg_k_k_id
        WHERE bwgk.bwg_k_bwg_id = ?
        ORDER BY k.k_name
    """, (bwg_id,))
    for name, stufe in cur.fetchall():
        ws.cell(row, 1, f"{name}: {stufe}")
        row += 1
    row += 1

    # Interessen
    row = add_section(ws, "Interessen", row)
    cur.execute("""
        SELECT i.i_name
        FROM tbl_bwg_i bwgi
        JOIN tbl_interessen i ON i.i_id = bwgi.bwg_i_i_id
        WHERE bwgi.bwg_i_bwg_id = ?
        ORDER BY i.i_name
    """, (bwg_id,))
    for (name,) in cur.fetchall():
        ws.cell(row, 1, name)
        row += 1
    row += 2

    # Abschluss
    ws.cell(row, 1, f"Dortmund, {datetime.today():%d.%m.%Y}").alignment = align_left
    ws.cell(row, 2, "Unterschrift").alignment = align_center
    img2 = os.path.join(os.path.dirname(__file__), "unterschrift.png")
    if os.path.exists(img2):
        im2 = Image(img2)
        im2.width = 100
        im2.height = 50
        im2.anchor = f"B{row + 1}"
        ws.add_image(im2)

    for r in range(1, row + 10):
        ws.row_dimensions[r].height = 18

    auto_adjust_width(ws)
    fn = os.path.join(os.path.dirname(__file__), f"Lebenslauf_{bwg_id}_{datetime.today():%Y%m%d}.xlsx")
    wb.save(fn)
    messagebox.showinfo("Erfolg", f"Gespeichert als {os.path.basename(fn)}")
    conn.close()

def main():
    bewerbungen = get_bewerbungen()
    if not bewerbungen:
        messagebox.showinfo("Info", "Keine Bewerbungen gefunden.")
        return
    root = tk.Tk()
    root.title("Lebenslauf Export")
    ttk.Label(root, text="Bitte Bewerbung auswählen:").pack(padx=10, pady=5)
    combo = ttk.Combobox(root, values=[b[1] for b in bewerbungen], state="readonly", width=50)
    combo.pack(padx=10, pady=5)
    combo.current(0)
    def on_export():
        export_lebenslauf(bewerbungen[combo.current()][0])
    ttk.Button(root, text="Exportieren", command=on_export).pack(padx=10, pady=10)
    root.mainloop()

if __name__ == "__main__":
    main()
